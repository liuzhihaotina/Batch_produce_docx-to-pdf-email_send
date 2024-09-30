import smtplib
from email.message import EmailMessage
from email.headerregistry import Address, Group
import email.policy
import mimetypes
import base64
import pandas
import time
import traceback
import openpyxl
import os

class SendEmail(object):
    """
    python3 发送邮件类
    格式： html
    可发送多个附件
    """
    def __init__(self, smtp_server, smtp_user, smtp_passwd, sender, recipient):
        # 发送邮件服务器,常用smtp.163.com
        self.smtp_server = smtp_server
        # 发送邮件的账号
        self.smtp_user = smtp_user
        # 发送账号的客户端授权码
        self.smtp_passwd = smtp_passwd
        # 发件人
        self.sender = sender
        # 收件人
        self.recipient = recipient

        # Use utf-8 encoding for headers. SMTP servers must support the SMTPUTF8 extension
        # https://docs.python.org/3.6/library/email.policy.html
        self.msg = EmailMessage(email.policy.SMTPUTF8)

    def set_header_content(self, subject, content):
        """
        设置邮件头和内容
        :param subject: 邮件题头
        :param content: 邮件内容
        :return:
        """
        self.msg['From'] = self.sender
        self.msg['To'] = self.recipient
        self.msg['Subject'] = subject
        self.msg.set_content(content, subtype="html")

    def set_accessories(self, path_list: list):
        """
        添加附件
        :param path_list: [{"path": ""}, {"name": ""}]
        :return:
        """
        for path_dict in path_list:
            path = path_dict['path']
            name = path_dict['name']
            # print(path, name)
            ctype, encoding = mimetypes.guess_type(path)
            if ctype is None or encoding is not None:
                # No guess could be made, or the file is encoded (compressed), so
                # use a generic bag-of-bits type.
                ctype = 'application/octet-stream'

            maintype, subtype = ctype.split('/', 1)
            with open(path, 'rb') as fp:
                self.msg.add_attachment(fp.read(), maintype, subtype, filename=name)
                # self.msg.add_attachment(fp.read(), maintype, subtype, filename=self.dd_b64(name))

    def send_email(self):
        """
        发送邮件
        :return:
        """
        with smtplib.SMTP_SSL(self.smtp_server, port=465) as smtp:
            # HELO向服务器标志用户身份
            smtp.ehlo_or_helo_if_needed()
            # 登录邮箱服务器
            smtp.login(self.smtp_user, self.smtp_passwd)
            print("Email:{}==>{}".format(self.sender, self.recipient))
            smtp.send_message(self.msg)
            print("成功发送!")

    @staticmethod
    def dd_b64(param):
        """
        对邮件header及附件的文件名进行两次base64编码，防止outlook中乱码。
        email库源码中先对邮件进行一次base64解码然后组装邮件
        :param param: 需要防止乱码的参数
        :return:
        """
        param = '=?utf-8?b?' + base64.b64encode(param.encode('UTF-8')).decode() + '?='
        param = '=?utf-8?b?' + base64.b64encode(param.encode('UTF-8')).decode() + '?='
        return param


if __name__ == '__main__':
    dir_path = os.path.dirname(os.path.abspath(__file__))  # 获取 绝对路径主目录
    # 学校邮箱
    # smtp_server = "smtp.hit.edu.cn"
    # smtp_user = "user1@hit.edu.cn"  # 发送邮件的账号
    # smtp_passwd = "XXXXXXXXX"          # 发送账号的客户端授权码
    # sender = Address("张老师", "user1", "hit.edu.cn")
    # 163邮箱
    # smtp_server = "smtp.163.com"
    # smtp_user = "123456@163.com"  # 发送邮件的账号
    # smtp_passwd = "xxxxxxxxx"          # 发送账号的客户端授权码
    # sender = Address("李老师", "123456", "163.com")
    # QQ邮箱
    smtp_server = "smtp.qq.com"
    smtp_user = "111222333@qq.com"  # 发送邮件的账号
    smtp_passwd = "xxxxxxxxx"  # 发送账号的客户端授权码
    sender = Address("王老师", "111222333", "qq.com")

    # majors = ['电气', '控制', '动力']
    # 手动更改专业学科名称
    majors = ['控制']
    # 日期
    t = time.localtime()
    content_pass = f"""
        <html>
            <p>某某同学你好，</p>
            <p style="text-indent:2em;">经考核，你被哈尔滨工业大学（深圳）机电工程与自动化学院拟接收为2025年硕士，请查收附件预接收函文件，待你获得本科所在学校推免资格并于全国推免服务系统完成注册后，此函正式生效，否则此函无效。</p>
            <p style="text-indent:2em;">请尽快加入机电学院预推免接收QQ群：xxxxxxxxx，并将群昵称修改为学科 - 姓名（机械 / 控制 / 电气 / 动力 - 姓名）。进群代表接受我院预推免名额并确认报考，如放弃名额，请单独发送邮件“学科 - 姓名 - 放弃预推免资格”至本邮箱，谢谢配合！</p>
            <p style="text-align: right">哈尔滨工业大学（深圳）机电工程与自动化学院</p>
            <p style="text-align: right">{t.tm_year}年{t.tm_mon}月{t.tm_mday}日</p>
        </html>
    """
    # content_fail = """
    #     <html>
    #         <p>某某同学：</p>
    #         <p>您好，您参加哈尔滨工业大学（深圳）预推免复试，很遗憾地通知您，
    #         经材料审核与考核，未通过本次推免面试，感谢您选择哈尔滨工业大学（深圳），祝您前程似锦，一切顺利。</p>
    #         <p style="text-align: right">哈尔滨工业大学（深圳）机电工程与自动化学院</p>
    #         <p style="text-align: right">2024年8月2日</p>
    #     </html>
    # """
    # 邮件标题
    subject = "哈尔滨工业大学（深圳）机电工程与自动化学院推免生预接收通知"
    g = 0 # 统计还剩多少没发送
    major='控制'
    # for major in majors:
    print(f"------------------------------------{major}")
    df = pandas.read_excel(f'./邮箱表格/{major}学科预录取函.xlsx')
    # 获取发送次数数据 先打开我们的目标表格，再打开我们的目标表单
    wb = openpyxl.load_workbook(rf'./邮箱表格/{major}学科预录取函.xlsx')
    ws = wb['Sheet1']
    # 序号
    j = 1
    # print(len(df))
    for i in range(len(df)):
        j+=1
        stu_name = df.loc[i]['姓名']
        num=df.loc[i]['序号']
        # print(df.loc[i]['电子邮箱'], type(df.loc[i]['电子邮箱']))
        stu_email_pre, stu_email_domain = df.loc[i]['电子邮箱'].split('@')
        # print(stu_email_pre, stu_email_domain)
        # print(type(stu_email_pre), type(stu_email_domain))
        # 不发送直博生的，当时是后面单独发送的
        if df.loc[i]['学术型/专业学位'] != '直博生':
            pass_flag = True
        else:
            pass_flag = False
        if pass_flag:
            recipient = Group(addresses=[Address(stu_name, stu_email_pre, stu_email_domain)])
            path_list = [
                {"path": f"{dir_path}/生成文档/pdf/{major}/2025年哈工大（深圳）机电学院推免生预接收函-{stu_name}.pdf",
                "name":  f"2025年哈工大（深圳）机电学院推免生预接收函-{stu_name}.pdf"}]
            content = content_pass.replace('某某', stu_name)
        else:
            continue

        # 发送邮件
        sd = SendEmail(smtp_server, smtp_user, smtp_passwd, sender, recipient)  # 创建对象
        sd.set_header_content(subject, content)  # 设置题头和内容
        sd.set_accessories(path_list)  # 添加附件
        # print(path_list)
        # 如果发送次数为0，执行发送
        # （这是因为第一次发送可能由于发送方或者接受方的问题，发送失败；而已经有其他发送成功的情况）
        # 可以避免重复发送或者漏发
        if j*(int(df['发送次数'].values[i]==0)):
           try:
               sd.send_email()  # 发送邮件
               # 发送成功，统计次数+1
               ws.cell(row=i + 2, column=ws.max_column).value = df['发送次数'].values[i] + 1
           except smtplib.SMTPResponseException:
               # 不知为何有时候报这个类型的错，但不影响邮件发送，因此跳过
               # print(type(e))
               # 发送成功，统计次数+1
               ws.cell(row=i + 2, column=ws.max_column).value = df['发送次数'].values[i] + 1
           except:  # 其他报错正常执行后面的失败统计
               g += 1  # 统计发送失败的数量
               # 打印发送失败的同学姓名
               print(f"{stu_name}Error")
               # 打印错误，不过目前作者遇到的都是“None”，可以考虑注释改行
               print(traceback.print_exc())
               # 若有需要，可以取消下行注释；但不建议，因为发送失败了会跳出循环
               # 而是建议继续下一个循环，先把可以发送的全发送了
               # exit()
    wb.save(rf'./邮箱表格/{major}学科预录取函.xlsx')  # 相对路径
    # 打印还剩多少没发送，也就是执行一次后没能发送成功的数量
    print(f'还剩{g}没发送')
