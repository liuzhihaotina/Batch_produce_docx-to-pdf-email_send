# 额外添加一列，用于统计发送次数
import pandas
import openpyxl
# # '''
# # distance_list是一个列表，我们的目标是将该列表作为一列插入表格
# # '''
major = '控制'  #专业
df = pandas.read_excel(f'./邮箱表格/{major}学科预录取函.xlsx')
num=len(df)
distance_list=[0]*num
# # 先打开我们的目标表格，再打开我们的目标表单
wb=openpyxl.load_workbook(rf'./邮箱表格/{major}学科预录取函.xlsx')
ws = wb['Sheet1']
c=ws.max_column+1  # 新增一列，存储发送次数
# # 取出distance_list列表中的每一个元素，openpyxl的行列号是从1开始取得，所以我这里i从1开始取
ws.cell(row = 1, column = c).value='发送次数'
for i in range(1,len(distance_list)+1):
    # 写入位置的行列号可以任意改变，这里从第2行开始按行依次插入对应列
    ws.cell(row = i+1, column = c).value =0
# # 保存操作
wb.save(rf'./邮箱表格/{major}学科预录取函.xlsx')

